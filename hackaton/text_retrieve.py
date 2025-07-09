"""
Read Excel (.xlsx) file line by line.
"""
from openpyxl import load_workbook
import embedding
from opensearchpy import OpenSearch

def retrieve_text(file_path = "ResponseData.xlsx", sheet_name="GEICO"):
    """
    Read Excel file line by line (row by row).
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the sheet to read. If None, reads the active sheet.
    
    Yields:
        list: Each row as a list of cell values
    """
    try:
        # Load the workbook
        workbook = load_workbook(filename=file_path, read_only=True)
        
        # Select the worksheet
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                print(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
                return
        else:
            worksheet = workbook.active
        
        print(f"Reading from sheet: {worksheet.title}")
        print(f"Total rows: {worksheet.max_row}")
        print(f"Total columns: {worksheet.max_column}")
        print("-" * 50)
        
        # Read each row, skipping the first row (header)
        for row_num, row in enumerate(worksheet.iter_rows(values_only=True, min_row=2), 2):
            yield row_num, list(row)
            
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
    finally:
        if 'workbook' in locals():
            workbook.close()

def get_opensearch_client():
    client = OpenSearch(
        hosts=[{"host": "localhost", "port": 9200}],
        http_auth=None,
        use_ssl=False,
        verify_certs=False,
    )
    return client

if __name__ == "__main__":
    opensearch_client = get_opensearch_client()
    for row_num, row in retrieve_text("ResponseData.xlsx", "GEICO"):
        #print(f"Row {row_num}: {row}")
        context = f"{row[0]},{row[1]},request_id:{row[3]}, question:{row[4]},response:{row[5]}" 
        vector = embedding.generate_embedding(context)
        doc = {
            "content": context,
            "embedding": vector,
        }

        if vector and isinstance(vector, list) and len(vector) == 1 and isinstance(vector[0], list):
            vector = vector[0]  # Flatten the embedding if it's a single-element list
        if not vector or not isinstance(vector, list) or not all(isinstance(x, (int, float)) for x in vector):
            print(f"Skipping indexing due to empty or invalid embedding: {vector}")
            continue

        opensearch_client.index(index="document", id=row[0], body=doc)
        #print(vector)
        if row_num >= 20:  # Limit to first 10 rows for demonstration
            break